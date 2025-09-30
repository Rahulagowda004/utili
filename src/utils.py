import pandas as pd
import openpyxl
from pathlib import Path
import shutil

def drop_col(df):
    required_columns = [
        "Parent summary",
        "Summary",
        "Assignee",
        "Custom field (Start date)",
        "Due date",
        "Σ Original Estimate",
        "Time Spent",
        "Status"
    ]
    df = df[required_columns]
    return df

def rename_col(df):
    rename_dict = {
        "Parent summary": "Module / Category",
        "Summary": "Task Details",
        "Assignee": "Assigned To",
        "Custom field (Start date)": "Start Date",
        "Due date": "Due Date",
        "Σ Original Estimate": "Planned (Hrs)",
        "Time Spent": "Actual (Hrs)",
        "Status": "Status"
    }
    df = df.rename(columns=rename_dict)
    df["End Date"] = df["Due Date"]
    df["Risk / Comments / Comp Off"] = ""
    return df

def date_format(df):
    df["Start Date"] = pd.to_datetime(df["Start Date"], dayfirst=True, format="mixed").dt.strftime("%#d/%#m/%Y")
    df["Due Date"] = pd.to_datetime(df["Due Date"], dayfirst=True, format="mixed").dt.strftime("%#d/%#m/%Y")
    df["End Date"] = df["Due Date"]
    return df

def time_format(df):
    df["Planned (Hrs)"] = (df["Planned (Hrs)"] / 3600).round(2)
    df["Actual (Hrs)"] = (df["Actual (Hrs)"] / 3600).round(2)
    return df

def save_to_excel(df):
    # Define source template path
    src = Path(df) if isinstance(df, str) and Path(df).exists() else Path("artifacts/template.xlsx")
    
    if not src.exists():
        raise FileNotFoundError(f"Source file not found: {src}. Run pipeline(...) or check artifacts folder.")
    
    # Define destination path
    dst = Path("output/result.xlsx")
    dst.parent.mkdir(parents=True, exist_ok=True)
    
    # Copy template to destination
    shutil.copy2(src, dst)
    
    # Load the workbook
    wb = openpyxl.load_workbook(dst)
    ws = wb.active  # or wb['SheetName'] if you want a specific sheet
    
    # Clear existing data (optional, if you want to overwrite from row 2)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    
    # Write DataFrame to Excel starting from row 2 (assuming headers are in row 1)
    for r_idx, row in enumerate(df.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Save the workbook
    wb.save(dst)
    
    return str(dst)

def pipeline(df):
    df = drop_col(df)
    df = rename_col(df)
    df = date_format(df)
    df = time_format(df)
    path = save_to_excel(df)
    return path


# LHS                  RHS

# Module / Category = Parent summary
# Task Details = Summary
# Assigned To = Assignee
# Start Date = Custom field (Start date)
# Due Date = Due date
# Planned (Hrs) = Σ Original Estimate
# Actual (Hrs) = Time Spent
# Status = Status
# End Date = Due date
# Risk / Comments / Comp Off = .